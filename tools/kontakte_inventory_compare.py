#!/usr/bin/env python3
"""
Compare inventory rows, detect duplicates, and mark the latest row.
Outputs: *_report.xlsx with fills (kept / older duplicate / unique).

One CSV (no true "second sheet") -> duplicate check is WITHIN the file
by key + optional date column. To compare two sources, use two file paths
or an .xlsx with two sheet names.

  pip install openpyxl

Examples:
  python tools/kontakte_inventory_compare.py ~/Downloads/sp_kontakte_inventory_correct.csv
  python tools/kontakte_inventory_compare.py a.csv b.csv
  python tools/kontakte_inventory_compare.py data.xlsx --sheets "Alt" "Neu" \\
      --key "E-Mail" --date "Geändert"
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional


def eprint(*args: object) -> None:
    print(*args, file=sys.stderr)


@dataclass
class Row:
    data: dict[str, Any]
    source: str
    line_or_row: int


@dataclass
class DupGroup:
    key: str
    rows: list[Row] = field(default_factory=list)
    best_idx: int = 0


def _norm(s: object) -> str:
    if s is None:
        return ""
    return str(s).strip()


def _try_parse_date(v: object) -> Optional[datetime]:
    s = _norm(v)
    if not s:
        return None
    s = s.replace("Z", "+00:00")
    fmts = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%Y-%m-%d",
    )
    for f in fmts:
        try:
            return datetime.strptime(s[: min(len(s), 25)], f)
        except ValueError:
            pass
    try:
        d = date.fromisoformat(s[:10])
        return datetime(d.year, d.month, d.day)
    except ValueError:
        return None


def _row_sort_time(r: Row, col: Optional[str]) -> float:
    if not col or col not in r.data:
        return float("-inf")
    dt = _try_parse_date(r.data.get(col))
    if dt:
        return dt.timestamp()
    return float("-inf")


def load_csv(
    path: Path,
    label: str,
) -> tuple[list[str], list[Row]]:
    encs = ("utf-8-sig", "utf-8", "cp1252", "iso-8859-1")
    last_err: Optional[Exception] = None
    for enc in encs:
        try:
            with path.open("r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    raise ValueError("CSV has no header row")
                fieldnames = list(reader.fieldnames)
                rows: list[Row] = []
                for i, row in enumerate(reader, start=2):
                    rows.append(Row(data=dict(row), source=label, line_or_row=i))
                return fieldnames, rows
        except Exception as e:
            last_err = e
    raise RuntimeError(f"Failed to read {path}: {last_err}")


def load_xlsx_sheets(
    path: Path,
    sheet1: str,
    sheet2: str,
) -> tuple[list[str], list[Row]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit("Install openpyxl: pip install openpyxl") from e

    wb = load_workbook(path, read_only=True, data_only=True)
    for name in (sheet1, sheet2):
        if name not in wb.sheetnames:
            eprint(
                f"Available sheets: {wb.sheetnames!r} — {name!r} not found"
            )
            raise SystemExit(1)

    def read_sheet(
        sname: str, tag: str, header_row: int = 1
    ) -> tuple[list[str], list[Row]]:
        ws = wb[sname]
        it = iter(ws.iter_rows(min_row=header_row, values_only=True))
        header = next(it, None)
        if not header or not any(h is not None and str(h).strip() for h in header):
            return [], []
        fieldnames = [str(c).strip() if c is not None else "" for c in header]
        rows: list[Row] = []
        rno = header_row + 1
        for t in it:
            d = {fieldnames[i]: t[i] for i in range(len(fieldnames)) if i < len(t)}
            if not any(_norm(v) for v in d.values()):
                rno += 1
                continue
            rows.append(Row(data=d, source=tag, line_or_row=rno))
            rno += 1
        return fieldnames, rows

    a_fields, a_rows = read_sheet(sheet1, f"sheet:{sheet1}")
    b_fields, b_rows = read_sheet(sheet2, f"sheet:{sheet2}")
    for r in b_rows:
        a_rows.append(r)
    all_fields: list[str] = []
    seen: set[str] = set()
    for h in a_fields + b_fields:
        if h and h not in seen:
            seen.add(h)
            all_fields.append(h)
    return all_fields, a_rows


def guess_key_column(fieldnames: list[str], key_arg: Optional[str]) -> str:
    if key_arg:
        c = _match_col(key_arg, fieldnames)
        if c:
            return c
        eprint(f"Warning: --key {key_arg!r} not in columns; using guess.")
    lower = {f.lower(): f for f in fieldnames}
    for w in (
        "e-mail",
        "e_mail",
        "email",
        "mail",
        "e-mailadresse",
        "primary email",
    ):
        for lf, real in lower.items():
            if w in lf or lf == w:
                return real
    if "vorname" in lower and "nachname" in lower:
        return "@@vor_nach"
    if fieldnames:
        return fieldnames[0]
    raise SystemExit("No columns found — empty file?")


def _match_col(name: str, fieldnames: list[str]) -> str:
    n = name.strip().lower()
    for f in fieldnames:
        if f.strip().lower() == n:
            return f
    for f in fieldnames:
        if n in f.lower():
            return f
    return ""


def row_key(
    r: Row,
    key_col: str,
) -> str:
    if key_col == "@@vor_nach":
        a = _norm(
            r.data.get(
                _match_col("Vorname", list(r.data.keys())) or "Vorname", ""
            )
        )
        b = _norm(
            r.data.get(
                _match_col("Nachname", list(r.data.keys())) or "Nachname", ""
            )
        )
        return f"{a}\t{b}"
    if key_col not in r.data:
        return _norm(r.data.get(list(r.data.keys())[0], ""))
    return _norm(r.data.get(key_col))


def guess_date_column(
    fieldnames: list[str],
    key_col: str,
) -> Optional[str]:
    pat = re.compile(
        r"(modify|geänd|aktual|updated?|anlage|erstellt|erstell|datum|date|last|zeit)",
        re.I,
    )
    for f in fieldnames:
        if f == key_col:
            continue
        if pat.search(f or ""):
            return f
    return None


def run(
    paths: list[Path],
    key_arg: Optional[str],
    date_arg: Optional[str],
    xlsx_sheets: Optional[tuple[str, str]],
) -> int:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        eprint("Install: pip install openpyxl")
        return 1

    if xlsx_sheets and len(paths) == 1 and paths[0].suffix.lower() in (
        ".xlsx",
        ".xlsm",
    ):
        fieldnames, rows = load_xlsx_sheets(
            paths[0], xlsx_sheets[0], xlsx_sheets[1]
        )
    elif len(paths) == 1:
        fieldnames, rows = load_csv(paths[0], paths[0].name)
    elif len(paths) == 2:
        f1, r1 = load_csv(paths[0], paths[0].name)
        f2, r2 = load_csv(paths[1], paths[1].name)
        for r in r2:
            r1.append(r)
        fieldnames, rows = f1, r1
        for h in f2:
            if h and h not in fieldnames:
                fieldnames.append(h)
    else:
        eprint("Use one .csv, two .csv files, or one .xlsx with --sheets.")
        return 1

    if not fieldnames or not rows:
        eprint("No data rows to compare.")
        return 1

    key_col = guess_key_column(fieldnames, key_arg)
    date_col = _match_col(date_arg, fieldnames) if date_arg else None
    if not date_col:
        date_col = guess_date_column(fieldnames, key_col)

    groups: dict[str, DupGroup] = {}
    for r in rows:
        k = row_key(r, key_col)
        if not k:
            k = f"__empty_{id(r)}"
        if k not in groups:
            groups[k] = DupGroup(key=k, rows=[])
        groups[k].rows.append(r)

    for g in groups.values():
        if len(g.rows) == 1:
            g.best_idx = 0
            continue
        g.best_idx = max(
            range(len(g.rows)),
            key=lambda i: (
                _row_sort_time(g.rows[i], date_col),
                float(g.rows[i].line_or_row),
            ),
        )

    fill_keeper = PatternFill("solid", fgColor="C6EFCE")
    fill_dup_older = PatternFill("solid", fgColor="FFEB9C")
    fill_only = PatternFill("solid", fgColor="D9D9D9")
    font_header = Font(bold=True)
    al_wrap = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()
    ws = wb.active
    if ws is None:
        eprint("Failed to create workbook")
        return 1
    ws.title = "Compare"

    meta = [
        ("Key column", key_col),
        ("Date column (for latest)", date_col or "(none: row number used as tie-break)"),
        (
            "Rows (total / duplicate keys / unique keys)",
            f"{len(rows)} / {sum(1 for g in groups.values() if len(g.rows) > 1)} / {len(groups)}",
        ),
    ]
    for i, (a, b) in enumerate(meta, 1):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
        ws.cell(i, 1).font = font_header

    start = len(meta) + 2
    out_cols = (
        ["_status", "_is_latest", "_source", "_row"]
        + [c for c in fieldnames if c in rows[0].data or c]
    )
    for j, c in enumerate(out_cols, 1):
        cell = ws.cell(start, j, c)
        cell.font = font_header
        cell.fill = PatternFill("solid", fgColor="B4C6E7")

    def status_for(
        g: DupGroup, idx: int, r: Row
    ) -> tuple[str, str, str]:
        """Returns (status, is_latest, fill) where fill is keeper|older|unique."""
        n = len(g.rows)
        if n == 1:
            return ("unique", "n/a", "unique")
        if idx == g.best_idx:
            return (f"duplicate (latest of {n})", "yes", "keeper")
        return (f"duplicate (older, {n} total)", "no", "older")

    rr = start + 1
    for g in sorted(
        groups.values(), key=lambda x: (len(x.rows) == 1, -len(x.rows), x.key)
    ):
        for idx, r in enumerate(g.rows):
            st, is_lat, which = status_for(g, idx, r)
            vals = [st, is_lat, r.source, r.line_or_row]
            for c in out_cols[4:]:
                vals.append(r.data.get(c, ""))
            for cj, v in enumerate(vals, 1):
                cell = ws.cell(rr, cj, v)
                cell.alignment = al_wrap
                if which == "keeper":
                    cell.fill = fill_keeper
                elif which == "older":
                    cell.fill = fill_dup_older
                else:
                    cell.fill = fill_only
            rr += 1

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Key"
    summary["B1"] = "Count"
    summary["C1"] = "Note"
    summary["A1"].font = font_header
    summary["B1"].font = font_header
    summary["C1"].font = font_header
    sr = 2
    for g in sorted(groups.values(), key=lambda u: -len(u.rows)):
        if len(g.rows) < 2:
            continue
        summary.cell(sr, 1, g.key)
        summary.cell(sr, 2, len(g.rows))
        w = g.rows[g.best_idx]
        summary.cell(
            sr, 3, f"Latest: {w.source} row {w.line_or_row} " + (f"({date_col})" if date_col else "")
        )
        sr += 1
    if sr == 2:
        summary["A2"] = "(no duplicate keys)"

    out = paths[0].parent / f"{paths[0].stem}_duplicates_report.xlsx"
    wb.save(out)
    eprint(
        f"Wrote: {out}\n"
        f"Key={key_col!r}, date={date_col!r} — green=kept (latest), yellow=older duplicate, gray=unique"
    )
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "paths",
        nargs="+",
        type=Path,
        help="One CSV, two CSVs, or one .xlsx",
    )
    p.add_argument(
        "--key",
        default=None,
        help="Key column (default: E-Mail, or Vorname+Nachname, or first column)",
    )
    p.add_argument(
        "--date",
        default=None,
        help="Date column to choose latest (auto-guessed if possible)",
    )
    p.add_argument(
        "--sheets",
        nargs=2,
        metavar=("SHEET1", "SHEET2"),
        help="For .xlsx: two sheet names to merge and compare",
    )
    args = p.parse_args()
    sheets = (args.sheets[0], args.sheets[1]) if args.sheets else None
    return run(args.paths, args.key, args.date, sheets)


if __name__ == "__main__":
    raise SystemExit(main())
